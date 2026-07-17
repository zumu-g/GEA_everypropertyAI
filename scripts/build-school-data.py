#!/usr/bin/env python3
"""Build src/data/school-data.json: authoritative schools + government school
zones for the Casey/Cardinia region.

Sources (all open data):
  - Vic DET school zones (findmyschool.vic.gov.au polygons), CC-BY:
    https://discover.data.vic.gov.au/dataset/victorian-government-school-zones-2026
  - ACARA School Location + School Profile (type, sector, website, enrolments):
    https://acara.edu.au/contact-us/acara-data-access

Usage: python3 scripts/build-school-data.py [cache_dir]
Requires: openpyxl (pip install openpyxl). Re-run yearly when new zones drop.
"""

import io
import json
import os
import sys
import urllib.request
import zipfile

import openpyxl

ZONES_URL = "https://www.education.vic.gov.au/Documents/about/research/datavic/dv418_DataVic_School_Zones_2026_MAR26.zip"
ACARA_BASE = "https://dataandreporting.blob.core.windows.net/anrdataportal/Data-Access-Program"
PROFILE_URL = f"{ACARA_BASE}/School%20Profile%202025.xlsx"
LOCATION_URL = f"{ACARA_BASE}/School%20Location%202025.xlsx"

# Casey/Cardinia + margin (project scope is these two LGAs)
BBOX = {"min_lng": 145.05, "max_lng": 145.90, "min_lat": -38.40, "max_lat": -37.85}

OUT = os.path.join(os.path.dirname(__file__), "..", "src", "data", "school-data.json")


def fetch(url: str, cache_dir: str, name: str) -> str:
    path = os.path.join(cache_dir, name)
    if not os.path.exists(path):
        print(f"downloading {url}")
        urllib.request.urlretrieve(url, path)
    return path


def in_bbox(lng: float, lat: float) -> bool:
    return (BBOX["min_lng"] <= lng <= BBOX["max_lng"]
            and BBOX["min_lat"] <= lat <= BBOX["max_lat"])


def rings_in_bbox(geom) -> bool:
    polys = geom["coordinates"] if geom["type"] == "MultiPolygon" else [geom["coordinates"]]
    for poly in polys:
        for lng, lat in (pt[:2] for pt in poly[0]):
            if in_bbox(lng, lat):
                return True
    return False


def simplify(geom):
    """MultiPolygon/Polygon -> [[outer, hole...], ...] with 5dp coords."""
    polys = geom["coordinates"] if geom["type"] == "MultiPolygon" else [geom["coordinates"]]
    return [[[[round(pt[0], 5), round(pt[1], 5)] for pt in ring] for ring in poly] for poly in polys]


def load_zones(zip_path: str):
    zones = {"primary": [], "secondary": []}
    with zipfile.ZipFile(zip_path) as zf:
        for level, member in [
            ("primary", "Primary_Integrated_2026.geojson"),
            ("secondary", "Secondary_Integrated_Year7_2026.geojson"),
        ]:
            data = json.load(io.TextIOWrapper(zf.open(member), encoding="utf-8"))
            for f in data["features"]:
                if rings_in_bbox(f["geometry"]):
                    zones[level].append({
                        "school": f["properties"]["School_Name"],
                        "polygons": simplify(f["geometry"]),
                    })
    return zones


def rows(path: str, sheet_suffix: str):
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = next(ws for name in wb.sheetnames if name.endswith(sheet_suffix) for ws in [wb[name]])
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() for h in next(it)]
    for r in it:
        yield dict(zip(header, r))


def load_schools(location_path: str, profile_path: str):
    profiles = {}
    for r in rows(profile_path, "SchoolProfile 2025"):
        profiles[r["ACARA SML ID"]] = r

    schools = []
    for r in rows(location_path, "SchoolLocations 2025"):
        if r["State"] != "VIC" or r["Latitude"] is None:
            continue
        lat, lng = float(r["Latitude"]), float(r["Longitude"])
        if not in_bbox(lng, lat):
            continue
        p = profiles.get(r["ACARA SML ID"], {})
        sector = str(r["School Sector"]).lower()  # Government / Catholic / Independent
        stype = str(r["School Type"]).lower()     # Primary / Secondary / Combined / Special
        schools.append({
            "name": r["School Name"],
            "type": stype,
            "sector": sector,
            "yearRange": p.get("Year Range"),
            "students": p.get("Total Enrolments"),
            "website": p.get("School URL"),
            "lat": round(lat, 6),
            "lng": round(lng, 6),
        })
    return schools


def main():
    cache_dir = sys.argv[1] if len(sys.argv) > 1 else "/tmp/school-data-cache"
    os.makedirs(cache_dir, exist_ok=True)
    zones = load_zones(fetch(ZONES_URL, cache_dir, "zones.zip"))
    schools = load_schools(
        fetch(LOCATION_URL, cache_dir, "acara-location.xlsx"),
        fetch(PROFILE_URL, cache_dir, "acara-profile.xlsx"),
    )
    out = {
        "generated": "2026-07-17",
        "zoneYear": 2026,
        "bbox": BBOX,
        "schools": schools,
        "zones": zones,
    }
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(out, f, separators=(",", ":"))
    print(f"wrote {OUT}: {len(schools)} schools, "
          f"{len(zones['primary'])} primary zones, {len(zones['secondary'])} secondary zones, "
          f"{os.path.getsize(OUT) // 1024} KB")


if __name__ == "__main__":
    main()
